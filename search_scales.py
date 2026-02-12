import pandas as pd
import openpyxl

file_path = r"c:\Users\Marcelo Beltran\Documents\IDEA\GIT\AuditorOperacional\riskops-platform-1\docs\Matriz escala de controles 2025 VB _admtarifa_ejemplo.xlsm"

def find_text_in_all_sheets(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in workbook.sheetnames:
        print(f"\nSearching in sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_str = " ".join([str(x) for x in row if x is not None])
            if any(k in row_str for k in ['Oportunidad', 'Alcance', 'Tipo de control', 'Segregación']):
                print(row_str)

if __name__ == "__main__":
    find_text_in_all_sheets(file_path)
